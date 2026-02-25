#!/usr/bin/env python3
"""
Seed the database from the Excel file: Grants.xlsx
Run from backend directory: python scripts/seed_from_excel.py
"""
import asyncio
import sys
import os
import re

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import AsyncSessionLocal, init_db
from app.models.grant import Grant, GrantDeadline
from app.models.user import User
import bcrypt as _bcrypt
from datetime import datetime


def hash_password(password: str) -> str:
    return _bcrypt.hashpw(password.encode('utf-8'), _bcrypt.gensalt()).decode('utf-8')

EXCEL_PATH = os.environ.get(
    "EXCEL_PATH",
    os.path.join(os.path.expanduser("~"), "Desktop", "\u0413\u0440\u0430\u043d\u0442\u044b", "\u0413\u0440\u0430\u043d\u0442\u044b.xlsx")
)

CATEGORY_MAP = {
    "физические лица": "individual",
    "физ": "individual",
    "нко": "nko",
    "некоммерческ": "nko",
    "организац": "nko",
    "бизнес": "business",
    "ип": "business",
    "малый": "business",
}


def detect_category(text: str) -> str:
    if not text:
        return "individual"
    text_lower = text.lower()
    for keyword, category in CATEGORY_MAP.items():
        if keyword in text_lower:
            return category
    return "individual"


def parse_amount(text: str):
    if not text:
        return None, None
    text_str = str(text)
    numbers = re.findall(r'[\d\s]+(?:[.,]\d+)?', text_str.replace('\xa0', ' '))
    max_amount = None
    for num_str in numbers:
        try:
            clean = num_str.replace(" ", "").replace(",", "")
            val = int(float(clean))
            if val > 1000:
                max_amount = val
                break
        except Exception:
            pass
    return max_amount, text_str.strip()


async def seed():
    print(f"Loading Excel from: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "ГРАНТ" in name.upper():
            sheet_name = name
            break

    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    print(f"Using sheet: {sheet_name}")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("No data found!")
        return

    header = rows[0]
    print(f"Columns ({len(header)}): {[str(h)[:30] for h in header]}")

    await init_db()

    async with AsyncSessionLocal() as db:
        from sqlalchemy import select
        existing_admin = await db.execute(select(User).where(User.email == "admin@grants.local"))
        if not existing_admin.scalar_one_or_none():
            admin = User(
                email="admin@grants.local",
                password_hash=hash_password("admin123"),
                full_name="Администратор",
                is_admin=True,
                telegram_link_code="ADMIN001",
            )
            db.add(admin)
            print("Created admin user: admin@grants.local / admin123")

        existing = await db.execute(select(Grant))
        for g in existing.scalars().all():
            await db.delete(g)
        await db.commit()

        grants_added = 0
        for i, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue

            cols = list(row)
            while len(cols) < 10:
                cols.append(None)

            source_name = str(cols[0]).strip() if cols[0] else f"Grant #{i}"
            if source_name == "None" or not source_name:
                continue

            source_url = str(cols[1]).strip() if cols[1] else None
            if source_url == "None":
                source_url = None

            who_can_apply = str(cols[2]).strip() if cols[2] else None
            age_restrictions = str(cols[3]).strip() if cols[3] else None
            amount_raw = str(cols[4]).strip() if cols[4] else None
            window_schedule = str(cols[5]).strip() if cols[5] else None
            typical_docs = str(cols[6]).strip() if cols[6] else None
            reporting = str(cols[7]).strip() if cols[7] else None
            critical_notes = str(cols[8]).strip() if cols[8] else None
            submission_target = str(cols[9]).strip() if cols[9] else None

            max_amount, max_amount_text = parse_amount(amount_raw)
            category = detect_category(who_can_apply or "")

            grant = Grant(
                source_name=source_name,
                source_url=source_url,
                who_can_apply=who_can_apply if who_can_apply != "None" else None,
                age_restrictions=age_restrictions if age_restrictions != "None" else None,
                max_amount=max_amount,
                max_amount_text=max_amount_text,
                window_schedule=window_schedule if window_schedule != "None" else None,
                typical_docs=typical_docs if typical_docs != "None" else None,
                reporting=reporting if reporting != "None" else None,
                critical_notes=critical_notes if critical_notes != "None" else None,
                submission_target=submission_target if submission_target != "None" else None,
                category=category,
                is_active=True,
            )
            db.add(grant)
            grants_added += 1
            print(f"  + {source_name[:60]} [{category}] {max_amount_text or ''}")

        await db.commit()
        print(f"\nSeeded {grants_added} grants successfully!")

        result = await db.execute(select(Grant))
        all_grants = result.scalars().all()
        print(f"Total grants in DB: {len(all_grants)}")


if __name__ == "__main__":
    asyncio.run(seed())
